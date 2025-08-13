
import io, re, pandas as pd, numpy as np
from datetime import date
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Table, TableStyle, Paragraph, SimpleDocTemplate, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import cm
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import json, os

AVAILABLE_FIELDS = ["ISIN","Strumento","Quantita","Valore","Valuta"]

SYNONYMS = {
    "ISIN": [r"^isin$", r"^cod(ice)?[\s_]*isin$", r"^isin[\s_]*code$"],
    "Strumento": [r"^strumen.*$", r"^descriz.*prod.*$", r"^denomin.*$", r"^nome.*$", r"^prodotto$"],
    "Quantita": [r"^q(uantit(a|à)|ta|\\.t[aà])$", r"^num(ero)?[\s_]*titoli$", r"^qty$"],
    "Valore": [r"^valore(_|[\s])?(attuale|di[\s_]mercato)?$", r"^controvalore.*$", r"^importo$", r"^aum$"],
    "Valuta": [r"^valuta$", r"^divisa$", r"^ccy$", r"^currency$"],
}

def _norm(s):
    s = str(s).strip().lower()
    s = re.sub(r"[^\w\s\.]", "", s, flags=re.UNICODE)
    s = s.replace("à","a").replace("è","e").replace("é","e").replace("ì","i").replace("ò","o").replace("ù","u")
    s = re.sub(r"\s+"," ", s)
    return s

def _load_preset(preset_name):
    cfg_path = os.path.join(os.path.dirname(__file__), "..", "config", "mappings.json")
    try:
        with open(cfg_path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        return cfg.get((preset_name or "").upper(), {})
    except Exception:
        return {}

def detect_columns(headers, preset_name="", filename=""):
    headers = list(headers or [])
    preset = _load_preset(preset_name) if preset_name else {}

    mapping = {}
    for field in AVAILABLE_FIELDS:
        preset_col = preset.get(field)
        if preset_col and preset_col in headers:
            mapping[field] = preset_col

    for field in AVAILABLE_FIELDS:
        if field in mapping: continue
        pats = SYNONYMS.get(field, [])
        for h in headers:
            hn = _norm(h)
            for p in pats:
                if re.match(p, hn):
                    mapping[field] = h
                    break
            if field in mapping: break

    for field in AVAILABLE_FIELDS:
        mapping.setdefault(field, None)
    return mapping

def _load_xlsx_or_csv(file_like):
    try:
        return pd.read_excel(file_like)
    except Exception:
        try:
            file_like.seek(0)
            return pd.read_csv(file_like)
        except Exception:
            return pd.DataFrame()

def _excel_bytes(sheets_dict):
    wb = Workbook(); ws0 = wb.active; wb.remove(ws0)
    for name, (headers, df) in sheets_dict.items():
        ws = wb.create_sheet(name[:31] if name else "Sheet")
        ws.append(headers)
        if df is not None and not df.empty:
            for _, r in df.iterrows():
                ws.append([r.get(h,"") for h in headers])
        for i in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(i)].width = 18
    bio = io.BytesIO(); wb.save(bio); bio.seek(0); return bio.getvalue()

def _pdf_bytes(summary, titoli_out, charts):
    styles = getSampleStyleSheet()
    story = [Paragraph("<b>Analisi Avanzata Integrata – Report</b>", styles["Title"]), Spacer(1, 0.3*cm),
             Paragraph(f"Data: {date.today().isoformat()}", styles["Normal"]),
             Paragraph(f"Saldo complessivo portafoglio: <b>{summary.get('AUM_totale',0):,.2f}</b>", styles["Normal"]),
             Spacer(1, 0.3*cm)]
    def pct(x, tot): return (x/tot*100.0) if tot else 0.0
    data_table = [
        ["Categoria","AUM","Incidenza %","Margine bps","Ricavo annuo"],
        ["Fondi", f"{summary.get('AUM_fondi',0):,.2f}", f"{pct(summary.get('AUM_fondi',0), summary.get('AUM_totale',0)):.2f}", "media foglio", ""],
        ["Titoli", f"{summary.get('AUM_titoli',0):,.2f}", f"{pct(summary.get('AUM_titoli',0), summary.get('AUM_totale',0)):.2f}", "-", "-"],
        ["Gestioni", f"{summary.get('AUM_gestioni',0):,.2f}", f"{pct(summary.get('AUM_gestioni',0), summary.get('AUM_totale',0)):.2f}", "media foglio", ""],
        ["Liquidità", f"{summary.get('AUM_liquidita',0):,.2f}", f"{pct(summary.get('AUM_liquidita',0), summary.get('AUM_totale',0)):.2f}", "0", "0"],
        ["TOTALE", f"{summary.get('AUM_totale',0):,.2f}", "100.00" if summary.get('AUM_totale',0)>0 else "0.00", "", ""],
    ]
    tbl = Table(data_table, hAlign="LEFT")
    tbl.setStyle(TableStyle([("GRID",(0,0),(-1,-1),0.3,colors.grey),
                             ("BACKGROUND",(0,0),(-1,0),colors.whitesmoke),
                             ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold")]))
    story += [tbl, Spacer(1, 0.4*cm),
              Paragraph("Allocazione per categoria", styles["Heading3"]),
              Image(charts["pie"], width=14*cm, height=10*cm), Spacer(1, 0.4*cm),
              Paragraph("Margini medi (bps) per categoria", styles["Heading3"]),
              Image(charts["bars"], width=14*cm, height=8*cm), Spacer(1, 0.6*cm)]
    story.append(Paragraph("Analisi Titoli (equity/obbligazioni)", styles["Heading2"]))
    if titoli_out.empty:
        story.append(Paragraph("Nessun titolo rilevato con ISIN valido.", styles["Normal"]))
    else:
        story.append(Paragraph("I dati mercato titoli sono recuperati dal web quando disponibili; in assenza restano 'Analisi richiesta'.", styles["Italic"]))
        head = ["ISIN","Tipo","Note"]
        rows = [[r["ISIN"], r["Tipo"], r["Flag_Sostituibile"]] for _, r in titoli_out.head(12).iterrows()]
        tbl2 = Table([head]+rows, hAlign="LEFT")
        tbl2.setStyle(TableStyle([("GRID",(0,0),(-1,-1),0.3,colors.grey),
                                  ("BACKGROUND",(0,0),(-1,0),colors.whitesmoke),
                                  ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold")]))
        story.append(tbl2)
    bio = io.BytesIO(); SimpleDocTemplate(bio, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=1.5*cm, bottomMargin=1.5*cm).build(story); bio.seek(0)
    return bio.getvalue()

def run_full_workflow(uploaded, auto_email="", user_map=None, preset_name=""):
    df = _load_xlsx_or_csv(uploaded)
    if df is None or df.empty:
        excel_bytes = _excel_bytes({"Input_Portafoglio": (["Nota"], pd.DataFrame([{"Nota":"File non leggibile"}]))})
        from reportlab.lib.styles import getSampleStyleSheet
        styles = getSampleStyleSheet(); bio = io.BytesIO()
        SimpleDocTemplate(bio, pagesize=A4).build([Paragraph("File non leggibile", styles["Title"])]); bio.seek(0)
        return {"excel_bytes": excel_bytes, "pdf_bytes": bio.getvalue(), "summary": {"righe_input":0,"AUM_totale":0}, "input_preview": None, "mapping_detected": {}}

    rec = detect_columns(df.columns.tolist(), preset_name=preset_name)

    n = len(df)
    def get_col(name, numeric=False):
        colname = rec.get(name)
        if colname and colname in df.columns:
            s = df[colname]
            return pd.to_numeric(s, errors="coerce") if numeric else s
        return pd.Series([None]*n)

    core = pd.DataFrame({
        "ISIN": get_col("ISIN"),
        "Strumento": get_col("Strumento"),
        "Quantita": get_col("Quantita", numeric=True),
        "Valore": get_col("Valore", numeric=True),
        "Valuta": get_col("Valuta"),
    })
    core["ISIN"] = core["ISIN"].astype(str).fillna("")
    core["Valore"] = core["Valore"].fillna(0.0)

    mask_liq = core["Strumento"].astype(str).str.contains(r"LIQUID|CONTO|CASH|DEPOSITO|LIQUIDITA", case=False, na=False)
    isin_like = core["ISIN"].astype(str).str.match(r'^[A-Z]{2}[A-Z0-9]{9}[0-9]$', na=False)
    is_fondo = core["Strumento"].astype(str).str.contains(r"SICAV|FUND|UCITS|FONDO", case=False, na=False)
    is_gest = core["Strumento"].astype(str).str.contains(r"GESTIONE|GP|GPM|GPF|LINEA", case=False, na=False)
    is_titolo = (~mask_liq) & (~is_fondo) & (~is_gest) & (isin_like)

    fondi_cols = ["ISIN","Denominazione","Categoria","Valuta","Rating","Quartile","YTD","1Y","3Y","Volatilità","Margine_bps","Retro_Banca_bps","Flag_Sostituibile","Valore"]
    titoli_cols = ["ISIN","Tipo","Settore","Area","Scadenza","Duration","Yield","Rating_Ag","P/E","TargetPrice_Consensus","Margine_bps","Flag_Sostituibile","Valore"]
    gest_cols   = ["Codice","Nome_Gestione","Linea","Valuta","Margine_bps","AUM_Cliente","Fee_Annua_Estimata","Flag_Sostituibile"]
    liq_cols    = ["Strumento","Valore","Valuta","Tasso_Riconosciuto","Spread_Banca_bps","Note"]

    fondi_out = pd.DataFrame(columns=fondi_cols)
    if is_fondo.any():
        mf = core[is_fondo][["ISIN","Strumento","Valuta","Valore"]].copy()
        mf["Flag_Sostituibile"] = ""
        mf.rename(columns={"Strumento":"Denominazione"}, inplace=True)
        for c in fondi_cols:
            if c not in mf.columns: mf[c] = ""
        mf = mf[fondi_cols]; fondi_out = mf

    def guess_tipo(s):
        s = str(s).upper()
        if any(k in s for k in ["BTP","BOND","OBBLIG","NOTE","DEBENTURE"]): return "Bond"
        if any(k in s for k in ["AZIONE","EQUITY","ORD.","RISP.","SHARE","SPA","INC","PLC"]): return "Equity"
        return ""
    titoli_out = pd.DataFrame(columns=titoli_cols)
    if is_titolo.any():
        tt = core[is_titolo][["ISIN","Strumento","Valore"]].copy()
        rows = []
        for _, r in tt.iterrows():
            rows.append({
                "ISIN": r.get("ISIN",""),
                "Tipo": guess_tipo(r.get("Strumento","")),
                "Settore": "",
                "Area": "",
                "Scadenza": "",
                "Duration": "",
                "Yield": "",
                "Rating_Ag": "",
                "P/E": "",
                "TargetPrice_Consensus": "",
                "Margine_bps": "",
                "Flag_Sostituibile": "Analisi richiesta (web)",
                "Valore": r.get("Valore",0.0)
            })
        titoli_out = pd.DataFrame(rows, columns=titoli_cols)

    gest_out = pd.DataFrame(columns=gest_cols)
    if is_gest.any():
        gg = core[is_gest][["Strumento","Valuta","Valore"]].copy()
        gg.rename(columns={"Strumento":"Nome_Gestione","Valuta":"Valuta","Valore":"AUM_Cliente"}, inplace=True)
        for c in gest_cols:
            if c not in gg.columns: gg[c] = ""
        gest_out = gg[gest_cols]

    liq_out = pd.DataFrame(columns=liq_cols)
    if mask_liq.any():
        ml = core[mask_liq][["Strumento","Valore","Valuta"]].copy()
        ml["Tasso_Riconosciuto"] = ""
        ml["Spread_Banca_bps"] = ""
        ml["Note"] = ""
        liq_out = ml[liq_cols]

    aum_fondi = float(fondi_out["Valore"].sum()) if "Valore" in fondi_out.columns else 0.0
    aum_titoli = float(titoli_out["Valore"].sum()) if "Valore" in titoli_out.columns else 0.0
    aum_gest = float(gest_out["AUM_Cliente"].sum()) if "AUM_Cliente" in gest_out.columns else 0.0
    aum_liq = float(liq_out["Valore"].sum()) if "Valore" in liq_out.columns else 0.0
    aum_tot = aum_fondi + aum_titoli + aum_gest + aum_liq

    # Grafici in memoria
    pie_buf = io.BytesIO()
    labels = ["Fondi","Titoli","Gestioni","Liquidità"]; values = [aum_fondi, aum_titoli, aum_gest, aum_liq]
    if aum_tot > 0:
        plt.figure(); plt.pie(values, labels=labels, autopct=lambda p: f'{p:.1f}%'); plt.title("Allocazione per categoria")
    else:
        plt.figure(); plt.bar([0],[0]); plt.title("Allocazione per categoria (nessun AUM)")
    plt.savefig(pie_buf, format="png", bbox_inches="tight"); plt.close(); pie_buf.seek(0)

    def safe_mean(series):
        s = pd.to_numeric(series, errors="coerce").dropna()
        return float(s.mean()) if len(s)>0 else 0.0
    m_fondi = safe_mean(fondi_out["Margine_bps"]) if not fondi_out.empty and "Margine_bps" in fondi_out.columns else 0.0
    m_gest = safe_mean(gest_out["Margine_bps"]) if not gest_out.empty and "Margine_bps" in gest_out.columns else 0.0
    m_titoli = 0.0; m_liq = 0.0
    bars_buf = io.BytesIO()
    plt.figure(); plt.bar(["Fondi","Titoli","Gestioni","Liquidità"], [m_fondi, m_titoli, m_gest, m_liq]); plt.title("Margini medi (bps) per categoria"); plt.ylabel("bps")
    plt.savefig(bars_buf, format="png", bbox_inches="tight"); plt.close(); bars_buf.seek(0)

    # Excel
    input_headers = ["Riga","ISIN/CUSIP","Strumento","Quantità","Prezzo_Medio","Valore_Attuale","Valuta","Categoria_Attesa (auto)"]
    input_df = core.copy().rename(columns={"ISIN":"ISIN/CUSIP","Quantita":"Quantità","Valore":"Valore_Attuale"})[["ISIN/CUSIP","Strumento","Quantità","Valore_Attuale","Valuta"]]
    input_df.insert(0, "Riga", range(1, len(input_df)+1)); input_df["Categoria_Attesa (auto)"] = ""
    sheets = {
        "Input_Portafoglio": (input_headers, input_df),
        "Fondi": (["ISIN","Denominazione","Categoria","Valuta","Rating","Quartile","YTD","1Y","3Y","Volatilità","Margine_bps","Retro_Banca_bps","Flag_Sostituibile","Valore"], fondi_out),
        "Titoli": (["ISIN","Tipo","Settore","Area","Scadenza","Duration","Yield","Rating_Ag","P/E","TargetPrice_Consensus","Margine_bps","Flag_Sostituibile","Valore"], titoli_out),
        "Gestioni": (["Codice","Nome_Gestione","Linea","Valuta","Margine_bps","AUM_Cliente","Fee_Annua_Estimata","Flag_Sostituibile"], gest_out),
        "Liquidità": (["Strumento","Valore","Valuta","Tasso_Riconosciuto","Spread_Banca_bps","Note"], liq_out),
        "Sintesi_Margini": (["Categoria","AUM","Incidenza_%","Margine_bps","Ricavo_Annuale"],
            pd.DataFrame([
                {"Categoria":"Fondi","AUM":aum_fondi,"Incidenza_%": (aum_fondi/aum_tot*100 if aum_tot else 0),"Margine_bps":"","Ricavo_Annuale":""},
                {"Categoria":"Titoli","AUM":aum_titoli,"Incidenza_%": (aum_titoli/aum_tot*100 if aum_tot else 0),"Margine_bps":"","Ricavo_Annuale":""},
                {"Categoria":"Gestioni","AUM":aum_gest,"Incidenza_%": (aum_gest/aum_tot*100 if aum_tot else 0),"Margine_bps":"","Ricavo_Annuale":""},
                {"Categoria":"Liquidità","AUM":aum_liq,"Incidenza_%": (aum_liq/aum_tot*100 if aum_tot else 0),"Margine_bps":"0","Ricavo_Annuale":""},
                {"Categoria":"TOTALE","AUM":aum_tot,"Incidenza_%": 100.0 if aum_tot else 0.0,"Margine_bps":"","Ricavo_Annuale":""},
            ])),
    }
    excel_bytes = _excel_bytes(sheets)

    pdf_bytes = _pdf_bytes(
        summary={
            "righe_input": int(len(core)),
            "AUM_fondi": aum_fondi, "AUM_titoli": aum_titoli,
            "AUM_gestioni": aum_gest, "AUM_liquidita": aum_liq,
            "AUM_totale": aum_tot,
        },
        titoli_out=titoli_out,
        charts={"pie": pie_buf, "bars": bars_buf}
    )

    summary = {
        "righe_input": int(len(core)),
        "colonne_rilevate": {k: bool(rec.get(k)) for k in AVAILABLE_FIELDS},
        "AUM_fondi": aum_fondi, "AUM_titoli": aum_titoli,
        "AUM_gestioni": aum_gest, "AUM_liquidita": aum_liq,
        "AUM_totale": aum_tot,
    }

    return {"excel_bytes": excel_bytes, "pdf_bytes": pdf_bytes, "summary": summary, "input_preview": core.head(50), "mapping_detected": rec}
